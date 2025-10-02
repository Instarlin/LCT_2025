from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


def parse_job_xlsx(file_bytes: bytes) -> Dict[str, Any]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
    ]
    header_map = {header.lower(): index for index, header in enumerate(headers)}

    parsed_rows: List[Dict[str, Any]] = []
    for raw_values in worksheet.iter_rows(min_row=2, values_only=True):
        if not raw_values or all(value in (None, "") for value in raw_values):
            continue
        raw_row = {headers[i]: raw_values[i] for i in range(len(headers))}
        normalized = _normalize_row(raw_row, header_map)
        if normalized:
            parsed_rows.append(normalized)

    return {
        "parsed_at": datetime.utcnow().isoformat(),
        "summary": {
            "Количество исследований": len(parsed_rows),
        },
        "rows": parsed_rows,
    }


def _normalize_row(row: Dict[str, Any], header_map: Dict[str, int]) -> Dict[str, Any]:
    def _get(*keys: str) -> Any:
        for key in keys:
            lower = key.lower()
            if lower in header_map:
                original = list(row.keys())[header_map[lower]]
                return row.get(original)
        return None

    study_uid = _to_str(_get("study_uid"))
    series_uid = _to_str(_get("series_uid"))
    probability = _to_float(_get("probability_of_pathology", "probability_of_pathology_percent"))

    pathology_raw = _get("pathology")
    pathology = None
    if isinstance(pathology_raw, bool):
        pathology = pathology_raw
    elif isinstance(pathology_raw, str):
        pathology = pathology_raw.strip().lower() in {"true", "1", "yes", "да"}

    time_of_processing = _to_str(_get("time_of_processing", "processing_time", "time_of_analysis"))

    dangerous_type = _to_str(_get("most_dangerous_pathology_type"))
    hazard_type, hazard_probability = _extract_max_probability(row)
    if not dangerous_type and hazard_type:
        dangerous_type = hazard_type

    probability_of_anomaly = _to_float(
        _get("probability_of_anomaly", "probability_of_anomaly_percent", "anom_score")
    )
    if probability_of_anomaly is None:
        probability_of_anomaly = _extract_probability_by_prefix(row, "prob@anomaly")

    return {
        "study_uid": study_uid,
        "series_uid": series_uid,
        "probability_of_pathology": probability,
        "pathology": pathology,
        "time_of_processing": time_of_processing,
        "most_dangerous_pathology_type": dangerous_type,
        "hazard_probability": hazard_probability,
        "probability_of_anomaly": probability_of_anomaly,
    }


def _extract_max_probability(row: Dict[str, Any]) -> Tuple[str | None, float | None]:
    best_type = None
    best_value = None
    for key, value in row.items():
        if not isinstance(key, str):
            continue
        key_lower = key.lower()
        if key_lower.startswith("prob@"):
            numeric = _to_float(value)
            if numeric is not None and (best_value is None or numeric > best_value):
                best_value = numeric
                best_type = key_lower.replace("prob@", "")
    return best_type, best_value


def _extract_probability_by_prefix(row: Dict[str, Any], prefix: str) -> float | None:
    prefix_lower = prefix.lower()
    for key, value in row.items():
        if isinstance(key, str) and key.lower() == prefix_lower:
            return _to_float(value)
    return None


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(',', '.'))
    except (ValueError, TypeError):
        return None
